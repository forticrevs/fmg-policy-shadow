"""
Report generation for FMG Policy Shadow Analysis.

Produces HTML, Excel (openpyxl), and JSON outputs from RunResult data.
"""

import json
import os
from collections import OrderedDict
from datetime import datetime
from typing import Any, Dict, List, Optional, Set

from .models import (
    CanonicalPolicy,
    FindingType,
    PackageResult,
    PolicyAction,
    RunResult,
    ShadowFinding,
)

# Optional openpyxl import
try:
    import openpyxl
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SEVERITY_COLORS = {
    "CRITICAL": "#dc3545",
    "HIGH": "#fd7e14",
    "MEDIUM": "#ffc107",
    "LOW": "#0d6efd",
    "INFO": "#6c757d",
}

FINDING_TYPE_LABELS = {
    "full_conflict_shadow": "Full Conflict Shadow",
    "partial_conflict_shadow": "Partial Conflict Shadow",
    "full_redundant_coverage": "Full Redundant Coverage",
    "partial_redundant_overlap": "Partial Redundant Overlap",
    "indeterminate_due_to_unsupported_objects": "Indeterminate",
}

FINDING_TYPE_COLORS = {
    "full_conflict_shadow": "#dc3545",
    "partial_conflict_shadow": "#e8747c",
    "full_redundant_coverage": "#198754",
    "partial_redundant_overlap": "#0dcaf0",
    "indeterminate_due_to_unsupported_objects": "#adb5bd",
}

# Human-readable labels for a policy's origin in the evaluation order.
_SECTION_LABELS = {
    "local": "Policy Package",
    "global-header": "Global Header",
    "global-footer": "Global Footer",
}

_FULLY_SHADOWED_TYPES = {
    FindingType.FULL_CONFLICT_SHADOW,
    FindingType.FULL_REDUNDANT_COVERAGE,
}

_DISABLE_FINDING_LABELS = {
    FindingType.FULL_CONFLICT_SHADOW: "Fully Shadowed (Conflict)",
    FindingType.FULL_REDUNDANT_COVERAGE: "Fully Shadowed (Redundant)",
}


# ===================================================================
# JSON Report
# ===================================================================

def generate_json_report(run_result: RunResult, output_path: str) -> str:
    """Generate a machine-readable JSON report.

    Returns the output file path.
    """
    data: Dict[str, Any] = {
        "tool_version": run_result.tool_version,
        "run_timestamp": run_result.run_timestamp,
        "elapsed_seconds": run_result.elapsed_seconds,
        "summary_counts": run_result.summary_counts(),
        "package_results": [],
        "errors": run_result.errors,
    }

    for pr in run_result.package_results:
        pkg: Dict[str, Any] = {
            "fmg": pr.fmg,
            "adom": pr.adom,
            "package": pr.package,
            "total_policies": pr.total_policies,
            "global_header_policies": pr.global_header_policies,
            "global_footer_policies": pr.global_footer_policies,
            "effective_policies": pr.effective_policies,
            "elapsed_seconds": pr.elapsed_seconds,
            "finding_counts": {
                "full_conflict_shadow": pr.full_conflict_count,
                "partial_conflict_shadow": pr.partial_conflict_count,
                "full_redundant_coverage": pr.full_redundant_count,
                "partial_redundant_overlap": pr.partial_redundant_count,
                "indeterminate": pr.indeterminate_count,
                "total": len(pr.findings),
            },
            "findings": [f.to_dict() for f in pr.findings],
            "policies": _serialize_policies(pr.policies),
            "unsupported_objects": pr.unsupported_objects,
            "errors": pr.errors,
        }
        data["package_results"].append(pkg)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, default=str)

    return output_path


def _serialize_policies(policies: List[CanonicalPolicy]) -> List[dict]:
    """Serialize a list of CanonicalPolicy to dicts for JSON."""
    result = []
    for p in policies:
        raw = p.raw_data or {}
        result.append({
            "policyid": p.policyid,
            "name": p.name,
            "seq_num": p.seq_num,
            "policy_section": p.policy_section,
            "action": p.action.value if p.action else "",
            "status": p.status,
            "comments": p.comments,
            "package": p.package,
            "fmg": p.fmg,
            "adom": p.adom,
            "srcintf": p.srcintf.describe() if p.srcintf else "",
            "dstintf": p.dstintf.describe() if p.dstintf else "",
            "srcaddr": p.srcaddr.describe() if p.srcaddr else "",
            "dstaddr": p.dstaddr.describe() if p.dstaddr else "",
            "service": p.service.describe() if p.service else "",
            "schedule": p.schedule.describe() if p.schedule else "",
            "srcaddr_objects": raw.get("_raw_srcaddr", []),
            "dstaddr_objects": raw.get("_raw_dstaddr", []),
            "service_objects": raw.get("_raw_service", []),
            "schedule_objects": raw.get("_raw_schedule", []),
            "security_profiles": p.security_profiles,
            "has_unresolved": p.has_unresolved,
        })
    return result


# ===================================================================
# HTML Report
# ===================================================================

def generate_html_report(run_result: RunResult, output_path: str) -> str:
    """Generate a self-contained HTML report.

    Returns the output file path.
    """
    counts = run_result.summary_counts()
    timestamp = run_result.run_timestamp or datetime.now().isoformat()

    # Gather scope info
    fmgs = sorted(set(pr.fmg for pr in run_result.package_results))
    adoms = sorted(set(pr.adom for pr in run_result.package_results))
    packages = sorted(set(pr.package for pr in run_result.package_results))

    html_parts: List[str] = []
    html_parts.append(_html_head(timestamp))
    html_parts.append(_html_header(timestamp, run_result.tool_version))
    html_parts.append(_html_scope(fmgs, adoms, packages))
    html_parts.append(_html_dashboard(counts))
    html_parts.append(_html_package_cards(run_result.package_results))
    html_parts.append(_html_findings_detail(run_result.package_results))
    html_parts.append(
        _html_cli_script_builder(
            run_result.package_results,
            timestamp,
            run_result.tool_version,
        )
    )
    html_parts.append(_html_methodology())
    html_parts.append(_html_limitations())
    html_parts.append(_html_footer())

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(html_parts))

    return output_path


def _esc(text: Any) -> str:
    """Escape HTML special characters."""
    s = str(text) if text is not None else ""
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _policy_identity(policy: CanonicalPolicy):
    """Identity stable within one package result."""
    return (policy.policy_section, policy.seq_num, policy.policyid)


def _finding_target_identity(finding: ShadowFinding):
    """Map a finding back to the exact policy that was analyzed."""
    return (
        finding.shadowed_section,
        finding.shadowed_seq,
        finding.shadowed_policyid,
    )


def _policy_matches_package_result(
    policy: CanonicalPolicy,
    pr: PackageResult,
) -> bool:
    """Return whether a policy belongs to this exact report package."""
    return (
        policy.fmg == pr.fmg
        and policy.adom == pr.adom
        and policy.package == pr.package
    )


_MAX_FMG_POLICY_ID = 1071741824


def _validated_policy_id(value) -> Optional[int]:
    """Return a safe CLI policy ID, or None for ambiguous/invalid values."""
    if isinstance(value, bool):
        return None
    text = str(value).strip()
    if (
        not text
        or len(text) > 10
        or any(character not in "0123456789" for character in text)
    ):
        return None
    try:
        policy_id = int(text)
    except (ValueError, OverflowError):
        return None
    # Require an explicit, positive existing FortiManager policy ID.  ``edit``
    # can create an object, so defaults, section-title ID 0, and FMG's
    # reserved range must never reach remediation.
    if policy_id <= 0 or policy_id > _MAX_FMG_POLICY_ID:
        return None
    return policy_id


def _package_policy_id_counts(pr: PackageResult) -> Dict[int, int]:
    """Count valid IDs for rules defined directly in one policy package."""
    counts: Dict[int, int] = {}
    for policy in pr.policies:
        if not _policy_matches_package_result(policy, pr):
            continue
        if policy.policy_section != "local" or policy.is_section_title:
            continue
        policy_id = _validated_policy_id(policy.policyid)
        if policy_id is not None:
            counts[policy_id] = counts.get(policy_id, 0) + 1
    return counts


def _collect_disable_candidates(pr: PackageResult) -> List[Dict[str, Any]]:
    """Return rules the analyzer recommends for disabling.

    A finding is eligible only when a single, high-confidence relationship
    proves a rule defined in the policy package unreachable and the shadowing
    rule's install scope contains the target rule's scope. Composite-union
    findings are not automatic recommendations because the analyzer's
    composite coverage is intentionally heuristic.
    """
    policies_by_identity = {
        _policy_identity(policy): policy
        for policy in pr.policies
    }
    package_policy_id_counts = _package_policy_id_counts(pr)
    qualifying: OrderedDict[Any, List[ShadowFinding]] = OrderedDict()

    for finding in pr.findings:
        if (
            finding.fmg != pr.fmg
            or finding.adom != pr.adom
            or finding.package != pr.package
        ):
            continue
        if finding.finding_type not in _FULLY_SHADOWED_TYPES:
            continue
        if not finding.is_fully_unreachable:
            continue
        if finding.is_composite:
            continue
        if finding.confidence.value != "high":
            continue
        if finding.shadowed_section != "local":
            continue
        if (
            len(finding.shadowing_policyids) != 1
            or len(finding.shadowing_seqs) != 1
            or (
                finding.shadowing_sections
                and len(finding.shadowing_sections) != 1
            )
        ):
            continue

        target_key = _finding_target_identity(finding)
        target_policy = policies_by_identity.get(target_key)
        if target_policy is None:
            continue
        if not _policy_matches_package_result(target_policy, pr):
            continue
        if target_policy.status != "enable" or target_policy.is_section_title:
            continue
        if target_policy.action not in {
            PolicyAction.ACCEPT,
            PolicyAction.DENY,
        } or target_policy.has_unresolved:
            continue
        if (
            target_policy.srcaddr.unresolved_names
            or target_policy.dstaddr.unresolved_names
            or target_policy.service.unresolved_names
            or target_policy.schedule.unresolved_name
        ):
            continue
        policy_id = _validated_policy_id(target_policy.policyid)
        if policy_id is None:
            continue
        if package_policy_id_counts.get(policy_id) != 1:
            continue

        shadowing_sections = finding.shadowing_sections or ["local"]
        shadowing_key = (
            shadowing_sections[0],
            finding.shadowing_seqs[0],
            finding.shadowing_policyids[0],
        )
        shadowing_policy = policies_by_identity.get(shadowing_key)
        if shadowing_policy is None:
            continue
        if not _policy_matches_package_result(shadowing_policy, pr):
            continue
        # Only the package rule ID is emitted as an executable ``edit`` value.
        # Inherited global-header/footer policies legitimately use FMG's
        # reserved high ID ranges, so applying the package edit-ID limit to the
        # shadowing (comment-only) policy would reject valid relationships.
        if shadowing_policy.is_section_title:
            continue
        if shadowing_policy.seq_num >= target_policy.seq_num:
            continue
        # ``--include-disabled`` deliberately adds disabled policies to the
        # analysis, but they do not currently shadow active traffic and must
        # never qualify an active target for remediation.
        if shadowing_policy.status != "enable":
            continue
        if shadowing_policy.action not in {
            PolicyAction.ACCEPT,
            PolicyAction.DENY,
        } or shadowing_policy.has_unresolved:
            continue
        if (
            shadowing_policy.srcaddr.unresolved_names
            or shadowing_policy.dstaddr.unresolved_names
            or shadowing_policy.service.unresolved_names
            or shadowing_policy.schedule.unresolved_name
        ):
            continue
        # Schedule groups and cross-midnight recurring windows are handled
        # conservatively by the analyzer.  The remediation path stays simpler
        # and safer by accepting only package-default "always" schedules on
        # both sides of the proof.
        if (
            not target_policy.schedule.is_always
            or not shadowing_policy.schedule.is_always
        ):
            continue
        if (
            target_policy.srcaddr_negate
            or target_policy.dstaddr_negate
            or target_policy.service_negate
            or shadowing_policy.srcaddr_negate
            or shadowing_policy.dstaddr_negate
            or shadowing_policy.service_negate
        ):
            continue
        if not all((
            shadowing_policy.srcintf.contains(target_policy.srcintf),
            shadowing_policy.dstintf.contains(target_policy.dstintf),
            shadowing_policy.srcaddr.contains(target_policy.srcaddr),
            shadowing_policy.dstaddr.contains(target_policy.dstaddr),
            shadowing_policy.service.contains(target_policy.service),
            shadowing_policy.schedule.contains(target_policy.schedule),
        )):
            continue
        if not shadowing_policy.install_scope.contains(target_policy.install_scope):
            continue

        qualifying.setdefault(target_key, []).append(finding)

    candidates: List[Dict[str, Any]] = []
    for target_key, findings in qualifying.items():
        policy = policies_by_identity[target_key]
        labels = sorted({
            _DISABLE_FINDING_LABELS[finding.finding_type]
            for finding in findings
        })
        shadowing_ids = sorted({
            str(finding.shadowing_policyids[0])
            for finding in findings
        })
        candidates.append({
            "policy": policy,
            "policy_id": _validated_policy_id(policy.policyid),
            "selection_kind": "recommended",
            "selection_label": "Recommended",
            "script_selection_label": "Recommended by analyzer",
            "finding_labels": labels,
            "shadowing_policyids": shadowing_ids,
            "max_risk": max(finding.risk_score for finding in findings),
            "relationship_count": len(findings),
        })

    candidates.sort(
        key=lambda candidate: (
            candidate["policy"].seq_num,
            candidate["policy_id"],
        )
    )
    return candidates


def _manual_finding_label(finding: ShadowFinding) -> str:
    """Return a short finding label for the manual-selection table."""
    if finding.finding_type in _DISABLE_FINDING_LABELS:
        return _DISABLE_FINDING_LABELS[finding.finding_type]
    return FINDING_TYPE_LABELS.get(
        finding.finding_type.value,
        finding.finding_type.value,
    )


def _collect_manual_script_candidates(
    pr: PackageResult,
    recommended: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Return other enabled package rules available by manual choice.

    Manual selection deliberately bypasses the analyzer's recommendation
    criteria. It does not bypass CLI targeting constraints: the rule must be
    defined directly in this policy package, currently enabled, and have a
    unique policy ID that is safe to place in an ``edit`` command.
    """
    recommended_keys = {
        _policy_identity(candidate["policy"])
        for candidate in recommended
    }
    package_policy_id_counts = _package_policy_id_counts(pr)
    findings_by_target: Dict[Any, List[ShadowFinding]] = {}

    for finding in pr.findings:
        if (
            finding.fmg != pr.fmg
            or finding.adom != pr.adom
            or finding.package != pr.package
        ):
            continue
        key = _finding_target_identity(finding)
        findings_by_target.setdefault(key, []).append(finding)

    candidates: List[Dict[str, Any]] = []
    for policy in pr.policies:
        key = _policy_identity(policy)
        if key in recommended_keys:
            continue
        if not _policy_matches_package_result(policy, pr):
            continue
        if policy.policy_section != "local" or policy.is_section_title:
            continue
        if policy.status != "enable":
            continue
        policy_id = _validated_policy_id(policy.policyid)
        if (
            policy_id is None
            or package_policy_id_counts.get(policy_id) != 1
        ):
            continue

        findings = findings_by_target.get(key, [])
        labels = sorted({
            _manual_finding_label(finding)
            for finding in findings
        })
        shadowing_ids = sorted({
            str(policy_id)
            for finding in findings
            for policy_id in finding.shadowing_policyids
        })
        candidates.append({
            "policy": policy,
            "policy_id": policy_id,
            "selection_kind": "manual",
            "selection_label": "Manual choice",
            "script_selection_label": "Added manually",
            "finding_labels": labels or ["Not recommended automatically"],
            "shadowing_policyids": shadowing_ids,
            "max_risk": (
                max(finding.risk_score for finding in findings)
                if findings
                else None
            ),
            "relationship_count": len(findings),
        })

    candidates.sort(
        key=lambda candidate: (
            candidate["policy"].seq_num,
            candidate["policy_id"],
        )
    )
    return candidates


def _inherited_script_rule_count(pr: PackageResult) -> int:
    """Count enabled inherited rules that require a Global Database script."""
    return sum(
        1
        for policy in pr.policies
        if (
            _policy_matches_package_result(policy, pr)
            and policy.policy_section in {"global-header", "global-footer"}
            and not policy.is_section_title
            and policy.status == "enable"
        )
    )


def _severity_badge(severity: str) -> str:
    color = SEVERITY_COLORS.get(severity, "#6c757d")
    text_color = "#fff" if severity != "MEDIUM" else "#212529"
    return (
        f'<span style="display:inline-block;padding:2px 8px;border-radius:4px;'
        f'font-size:0.8em;font-weight:bold;color:{text_color};'
        f'background-color:{color}">{_esc(severity)}</span>'
    )


def _finding_type_badge(ftype: str) -> str:
    color = FINDING_TYPE_COLORS.get(ftype, "#adb5bd")
    label = FINDING_TYPE_LABELS.get(ftype, ftype)
    return (
        f'<span style="display:inline-block;padding:2px 8px;border-radius:4px;'
        f'font-size:0.8em;color:#fff;background-color:{color}">{_esc(label)}</span>'
    )


def _section_badge(section: str) -> str:
    """Small badge flagging a policy inherited from the global database.

    Returns an empty string for rules defined in the policy package so the
    common case stays uncluttered.
    """
    if section not in ("global-header", "global-footer"):
        return ""
    label = _SECTION_LABELS.get(section, section)
    return (
        f'<span style="display:inline-block;padding:1px 6px;border-radius:4px;'
        f'font-size:0.72em;font-weight:600;color:#fff;background-color:#6610f2;'
        f'margin-left:4px;">{_esc(label)}</span>'
    )


def _format_explanation(text: str) -> str:
    """Convert plain-text explanation into structured HTML.

    Handles a special SHADOWING_RULES: block (generated by composite
    explanations) by rendering it as a compact, scrollable table instead
    of dumping all policy references inline.
    """
    if not text:
        return '<div class="explanation-block"><p class="intro">No explanation available.</p></div>'

    # Dimension keywords to color-code
    _DIM_KEYWORDS = {
        "srcintf": "#0d6efd",
        "dstintf": "#0d6efd",
        "srcaddr": "#198754",
        "dstaddr": "#198754",
        "service": "#6f42c1",
        "schedule": "#fd7e14",
    }

    def _highlight(line: str) -> str:
        """Bold key terms and color-code dimension names in a line."""
        esc = _esc(line)
        # Bold key phrases
        for phrase in ("FULLY SHADOWED", "FULLY COVERED", "PARTIALLY SHADOWED",
                       "PARTIALLY OVERLAPS", "INDETERMINATE", "COMPOSITE UNION",
                       "NEVER be reached", "DIFFERENT action", "SAME action",
                       "UNREACHABLE", "redundant"):
            esc = esc.replace(_esc(phrase), f"<strong>{_esc(phrase)}</strong>")
        # Color-code dimension names
        for dim, color in _DIM_KEYWORDS.items():
            esc = esc.replace(dim, f'<span style="color:{color};font-weight:600">{dim}</span>')
        return esc

    # Separate the SHADOWING_RULES block if present
    rules_html = ""
    main_text = text
    if "\nSHADOWING_RULES:" in text:
        main_text, rules_raw = text.split("\nSHADOWING_RULES:", 1)
        rule_lines = [
            line.strip()
            for line in rules_raw.strip().split("\n")
            if line.strip()
        ]
        if rule_lines:
            rules_html = (
                '  <details class="rules-details" open>\n'
                '    <summary style="cursor:pointer;font-weight:600;margin:8px 0 4px 0;'
                'color:var(--text-color);font-size:0.9em;">'
                f'Shadowing Rules ({len(rule_lines)})</summary>\n'
                '    <div style="max-height:260px;overflow-y:auto;border:1px solid '
                'var(--border-color);border-radius:4px;margin-top:4px;">\n'
                '    <table style="width:100%;border-collapse:collapse;font-size:0.85em;">\n'
                '      <thead><tr style="position:sticky;top:0;'
                'background:var(--table-header-bg);z-index:1;">'
                '<th style="padding:4px 8px;text-align:left;border-bottom:1px solid '
                'var(--border-color);">Seq</th>'
                '<th style="padding:4px 8px;text-align:left;border-bottom:1px solid '
                'var(--border-color);">Policy ID</th>'
                '<th style="padding:4px 8px;text-align:left;border-bottom:1px solid '
                'var(--border-color);">Name</th>'
                '<th style="padding:4px 8px;text-align:left;border-bottom:1px solid '
                'var(--border-color);">Action</th>'
                '</tr></thead>\n<tbody>\n'
            )
            for rl in rule_lines:
                parts = rl.split("|", 3)
                if len(parts) == 4:
                    seq, pid, name, action = parts
                    action_color = "#198754" if action == "accept" else "#dc3545"
                    rules_html += (
                        f'<tr style="border-bottom:1px solid var(--border-color);">'
                        f'<td style="padding:3px 8px;">{_esc(seq)}</td>'
                        f'<td style="padding:3px 8px;font-family:monospace;">{_esc(pid)}</td>'
                        f'<td style="padding:3px 8px;">{_esc(name)}</td>'
                        f'<td style="padding:3px 8px;color:{action_color};'
                        f'font-weight:600;">{_esc(action)}</td></tr>\n'
                    )
            rules_html += '</tbody></table>\n</div>\n</details>\n'

    # Split into intro paragraph and bullet-point details
    parts = main_text.split("\n", 1)
    intro = parts[0].strip()
    html = '<div class="explanation-block">\n'
    html += f'  <p class="intro">{_highlight(intro)}</p>\n'

    if len(parts) > 1 and parts[1].strip():
        lines = parts[1].strip().split("\n")
        bullets = [
            line.strip().lstrip("- ").strip()
            for line in lines
            if line.strip().startswith("-")
            or line.strip().startswith("  -")
        ]
        if bullets:
            html += "  <ul>\n"
            for b in bullets:
                html += f"    <li>{_highlight(b)}</li>\n"
            html += "  </ul>\n"

    if rules_html:
        html += rules_html

    html += "</div>"
    return html


def _html_head(timestamp: str) -> str:
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>FMG Policy Shadow Analysis Report - {_esc(timestamp)}</title>
<style>
  :root {{
    --bg-body: #f8f9fa;
    --text-color: #212529;
    --heading-color: #1a1a2e;
    --heading-color-2: #16213e;
    --accent-color: #0d6efd;
    --card-bg: #fff;
    --shadow-color: rgba(0,0,0,0.1);
    --table-header-bg: #e9ecef;
    --hover-bg: #f1f3f5;
    --summary-bg: #e9ecef;
    --summary-hover-bg: #dee2e6;
    --border-color: #dee2e6;
    --pre-bg: #f8f9fa;
    --methodology-bg: #eef;
    --limitations-bg: #fff3cd;
    --pill-bg: #e9ecef;
    --muted-text: #6c757d;
    --overlap-th-bg: #f0f0f0;
    --pkg-group-header-color: #16213e;
    --header-gradient-start: #1a1a2e;
    --header-gradient-end: #16213e;
    --header-subtitle-color: #adb5bd;
  }}
  body.dark {{
    --bg-body: #0d1117;
    --text-color: #e6edf3;
    --heading-color: #e6edf3;
    --heading-color-2: #adbac7;
    --accent-color: #58a6ff;
    --card-bg: #161b22;
    --shadow-color: rgba(0,0,0,0.4);
    --table-header-bg: #1a1a2e;
    --hover-bg: #1c2333;
    --summary-bg: #1a1a2e;
    --summary-hover-bg: #16213e;
    --border-color: #30363d;
    --pre-bg: #161b22;
    --methodology-bg: #1a1a2e;
    --limitations-bg: #2d2a1e;
    --pill-bg: #1a1a2e;
    --muted-text: #adbac7;
    --overlap-th-bg: #1a1a2e;
    --pkg-group-header-color: #adbac7;
    --header-gradient-start: #0d1117;
    --header-gradient-end: #161b22;
    --header-subtitle-color: #adbac7;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
         line-height: 1.6; color: var(--text-color); background: var(--bg-body); padding: 20px;
         transition: background 0.3s, color 0.3s; }}
  .container {{ max-width: 1200px; margin: 0 auto; }}
  h1 {{ color: var(--heading-color); margin-bottom: 5px; }}
  h2 {{ color: var(--heading-color-2); margin: 30px 0 15px 0; padding-bottom: 8px; border-bottom: 2px solid var(--accent-color); }}
  h3 {{ color: var(--heading-color); margin: 20px 0 10px 0; }}
  .header {{ background: linear-gradient(135deg, var(--header-gradient-start) 0%, var(--header-gradient-end) 100%);
             color: #fff; padding: 30px; border-radius: 8px; margin-bottom: 25px; position: relative; }}
  .header h1 {{ color: #fff; border: none; }}
  .header .subtitle {{ color: var(--header-subtitle-color); font-size: 0.9em; }}
  .card {{ background: var(--card-bg); border-radius: 8px; padding: 20px; margin-bottom: 20px;
           box-shadow: 0 1px 3px var(--shadow-color); transition: background 0.3s; }}
  .dashboard {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
                gap: 15px; margin-bottom: 25px; }}
  .stat-box {{ background: var(--card-bg); border-radius: 8px; padding: 18px; text-align: center;
               box-shadow: 0 1px 3px var(--shadow-color); transition: background 0.3s; }}
  .stat-box .value {{ font-size: 2em; font-weight: bold; }}
  .stat-box .label {{ font-size: 0.85em; color: var(--muted-text); }}
  table {{ width: 100%; border-collapse: collapse; margin: 10px 0; font-size: 0.9em; }}
  th, td {{ padding: 8px 12px; text-align: left; border-bottom: 1px solid var(--border-color); }}
  th {{ background: var(--table-header-bg); font-weight: 600; position: sticky; top: 0; }}
  tr:hover {{ background: var(--hover-bg); }}
  details {{ margin: 8px 0; }}
  summary {{ cursor: pointer; padding: 10px 14px; background: var(--summary-bg); border-radius: 6px;
             font-weight: 500; user-select: none; transition: background 0.3s; }}
  summary:hover {{ background: var(--summary-hover-bg); }}
  details[open] summary {{ border-radius: 6px 6px 0 0; }}
  details .detail-body {{ padding: 14px; border: 1px solid var(--border-color); border-top: none;
                          border-radius: 0 0 6px 6px; background: var(--card-bg); }}
  pre.explanation {{ white-space: pre-wrap; word-break: break-word; background: var(--pre-bg);
                     border: 1px solid var(--border-color); border-radius: 4px; padding: 10px 14px;
                     font-family: inherit; font-size: 0.93em; line-height: 1.6; margin: 4px 0 10px;
                     color: var(--text-color); }}
  .overlap-table th {{ background: var(--overlap-th-bg); width: 160px; }}
  .pkg-card {{ border-left: 4px solid var(--accent-color); }}
  .methodology {{ background: var(--methodology-bg); border-left: 4px solid var(--accent-color); padding: 15px; margin: 10px 0; border-radius: 4px; }}
  .limitations {{ background: var(--limitations-bg); border-left: 4px solid #ffc107; padding: 15px; margin: 10px 0; border-radius: 4px; }}
  .error-list {{ color: #dc3545; }}
  .scope-pills {{ display: flex; flex-wrap: wrap; gap: 6px; margin: 5px 0; }}
  .scope-pill {{ background: var(--pill-bg); padding: 3px 10px; border-radius: 12px; font-size: 0.85em; }}
  .fmg-group {{ margin: 20px 0; }}
  .fmg-group-header {{ background: linear-gradient(135deg, #1a1a2e, #16213e); color: #fff;
                        padding: 12px 18px; border-radius: 8px 8px 0 0; font-size: 1.1em; }}
  .pkg-group {{ border: 1px solid var(--border-color); border-top: none; padding: 15px; margin-bottom: 0; background: var(--card-bg); }}
  .pkg-group:last-child {{ border-radius: 0 0 8px 8px; }}
  .pkg-group-header {{ font-size: 1em; color: var(--pkg-group-header-color); margin: 0 0 10px 0; padding-bottom: 6px;
                        border-bottom: 1px solid var(--border-color); }}
  .skipped-list {{ background: var(--limitations-bg); border-left: 3px solid #ffc107; padding: 8px 14px;
                   margin: 8px 0; border-radius: 4px; font-size: 0.88em; }}
  .dark-toggle {{ position: absolute; top: 20px; right: 20px; background: rgba(255,255,255,0.15);
                   border: 1px solid rgba(255,255,255,0.3); color: #fff; border-radius: 50%;
                   width: 40px; height: 40px; font-size: 1.2em; cursor: pointer;
                   display: flex; align-items: center; justify-content: center;
                   transition: background 0.3s, transform 0.2s; backdrop-filter: blur(4px); }}
  .dark-toggle:hover {{ background: rgba(255,255,255,0.25); transform: scale(1.1); }}
  .explanation-block {{ margin: 4px 0 10px; }}
  .explanation-block .intro {{ font-size: 0.95em; line-height: 1.5; margin-bottom: 8px; padding: 8px 12px;
    background: var(--pre-bg, #f8f9fa); border-left: 3px solid var(--accent-color, #0d6efd); border-radius: 4px; }}
  .explanation-block ul {{ margin: 4px 0 0 20px; font-size: 0.9em; }}
  .explanation-block li {{ margin: 3px 0; color: var(--text-color, #212529); }}
  .risk-score {{ display: inline-block; padding: 2px 10px; border-radius: 4px; font-weight: bold; font-size: 0.9em; }}
  .risk-factors {{ font-size: 0.82em; color: var(--muted-text, #6c757d); margin-top: 2px; }}
  .remediation-intro {{ border-left: 4px solid #0d6efd; }}
  .remediation-warning {{ margin: 12px 0; padding: 10px 14px; border-left: 3px solid #fd7e14;
                          border-radius: 4px; background: var(--limitations-bg); }}
  .remediation-stats {{ display: flex; flex-wrap: wrap; gap: 8px; margin: 12px 0; }}
  .remediation-stats span {{ background: var(--pill-bg); border-radius: 14px; padding: 4px 11px;
                             font-size: 0.86em; }}
  .remediation-global-controls, .remediation-package-controls, .remediation-actions {{
    display: flex; align-items: center; flex-wrap: wrap; gap: 10px; margin: 12px 0;
  }}
  .remediation-package-controls {{ justify-content: space-between; }}
  .remediation-target {{ margin: 6px 0 12px; padding: 9px 12px; background: var(--pre-bg);
                         border-radius: 4px; }}
  .remediation-note {{ color: var(--muted-text); font-size: 0.86em; margin-top: 8px; }}
  .remediation-empty {{ margin-top: 12px; }}
  .remediation-table-wrap {{ overflow-x: auto; border: 1px solid var(--border-color);
                             border-radius: 5px; }}
  .remediation-table {{ margin: 0; min-width: 940px; }}
  .remediation-table .checkbox-column {{ width: 70px; text-align: center; }}
  .selection-badge {{ display: inline-block; border-radius: 12px; padding: 3px 9px;
                       font-size: 0.8em; font-weight: 600; white-space: nowrap; }}
  .selection-badge.recommended {{ color: #0f5132; background: #d1e7dd; }}
  .selection-badge.manual {{ color: #664d03; background: #fff3cd; }}
  .manual-policy-picker {{ margin-top: 16px; padding: 10px 12px;
                           border: 1px solid var(--border-color); border-radius: 5px; }}
  .manual-policy-picker summary {{ cursor: pointer; font-weight: 650; }}
  .manual-policy-picker > p {{ margin: 9px 0; color: var(--muted-text); }}
  .remediation-table input[type="checkbox"],
  .remediation-global-controls input[type="checkbox"],
  .remediation-package-controls input[type="checkbox"] {{
    width: 17px; height: 17px; vertical-align: middle; cursor: pointer;
  }}
  .comments-cell {{ max-width: 280px; white-space: normal; overflow-wrap: anywhere; }}
  .selection-status, .copy-status {{ color: var(--muted-text); font-size: 0.86em; }}
  .remediation-actions button, .remediation-global-controls button {{
    border: 0; border-radius: 5px; padding: 8px 13px; color: #fff; background: #0d6efd;
    font-weight: 600; cursor: pointer;
  }}
  .remediation-actions button.secondary-button,
  .remediation-global-controls button.secondary-button {{ background: #6c757d; }}
  .remediation-actions button:disabled {{ opacity: 0.45; cursor: not-allowed; }}
  .cli-preview {{ width: 100%; min-height: 300px; resize: vertical; padding: 12px;
                  border: 1px solid var(--border-color); border-radius: 5px;
                  color: var(--text-color); background: var(--pre-bg);
                  font: 0.86em/1.45 ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; }}
</style>
<script>
(function() {{
  try {{
    if (localStorage.getItem('fmg-dark-mode') === 'true') {{
      document.documentElement.style.background = '#0d1117';
    }}
  }} catch (error) {{
    // Some browsers restrict localStorage for file:// reports.
  }}
}})();
</script>
</head>
<body>
<div class="container">
"""


def _html_header(timestamp: str, version: str) -> str:
    return f"""<div class="header">
  <h1>FMG Policy Shadow Analysis Report</h1>
  <div class="subtitle">Generated: {_esc(timestamp)} &bull; Tool Version: {_esc(version)}</div>
  <button class="dark-toggle" id="darkToggle" title="Toggle dark mode" aria-label="Toggle dark mode">&#9790;</button>
</div>
"""


def _html_scope(fmgs: List[str], adoms: List[str], packages: List[str]) -> str:
    def _pills(items: List[str]) -> str:
        return "".join(f'<span class="scope-pill">{_esc(i)}</span>' for i in items)
    return f"""<div class="card">
  <h3>Analysis Scope</h3>
  <p><strong>FortiManagers ({len(fmgs)}):</strong></p>
  <div class="scope-pills">{_pills(fmgs)}</div>
  <p style="margin-top:8px"><strong>ADOMs ({len(adoms)}):</strong></p>
  <div class="scope-pills">{_pills(adoms)}</div>
  <p style="margin-top:8px"><strong>Packages ({len(packages)}):</strong></p>
  <div class="scope-pills">{_pills(packages)}</div>
</div>
"""


def _html_dashboard(counts: dict) -> str:
    items = [
        (str(counts.get("packages_analyzed", 0)), "Packages Analyzed", "#0d6efd"),
        (str(counts.get("total_policies", 0)), "Policies Analyzed", "#198754"),
        (str(counts.get("total_findings", 0)), "Total Findings", "#dc3545"),
        (str(counts.get("full_conflict_shadow", 0)), "Fully Shadowed (Conflict)", "#dc3545"),
        (str(counts.get("partial_conflict_shadow", 0)), "Partially Shadowed (Conflict)", "#fd7e14"),
        (str(counts.get("full_redundant_coverage", 0)), "Fully Shadowed (Redundant)", "#198754"),
        (str(counts.get("partial_redundant_overlap", 0)), "Partial Overlap (Redundant)", "#0dcaf0"),
        (str(counts.get("indeterminate", 0)), "Indeterminate (Unresolved)", "#6c757d"),
        (str(counts.get("errors", 0)), "Errors", "#dc3545"),
    ]
    boxes = ""
    for value, label, color in items:
        boxes += f'<div class="stat-box"><div class="value" style="color:{color}">{value}</div><div class="label">{label}</div></div>\n'
    return f"""<h2>Summary Dashboard</h2>
<div class="dashboard">
{boxes}
</div>
"""


def _html_package_cards(package_results: List[PackageResult]) -> str:
    if not package_results:
        return "<p>No packages analyzed.</p>"
    html = "<h2>Package Results</h2>\n"

    # Group by FMG instance
    fmg_groups: OrderedDict[str, List[PackageResult]] = OrderedDict()
    for pr in package_results:
        fmg_groups.setdefault(pr.fmg, []).append(pr)

    for fmg_host, prs in fmg_groups.items():
        html += f"<h3>FortiManager: {_esc(fmg_host)}</h3>\n"
        for pr in prs:
            findings_count = len(pr.findings)
            border_color = "#dc3545" if pr.full_conflict_count > 0 else (
                "#fd7e14" if pr.partial_conflict_count > 0 else "#0d6efd"
            )
            global_row = ""
            if pr.global_header_policies or pr.global_footer_policies:
                local_count = (
                    pr.total_policies
                    - pr.global_header_policies
                    - pr.global_footer_policies
                )
                global_row = (
                    f'    <tr><td>Global Header / Local / Footer</td>'
                    f'<td colspan="3"><strong>{pr.global_header_policies}</strong> global header '
                    f'+ <strong>{local_count}</strong> local '
                    f'+ <strong>{pr.global_footer_policies}</strong> global footer</td></tr>\n'
                )
            html += f"""<div class="card pkg-card" style="border-left-color:{border_color}">
  <h3>{_esc(pr.fmg)} / {_esc(pr.adom)} / {_esc(pr.package)}</h3>
  <table>
    <tr><td>Total Policies</td><td><strong>{pr.total_policies}</strong></td>
        <td>Effective Policies</td><td><strong>{pr.effective_policies}</strong></td></tr>
{global_row}    <tr><td>Fully Shadowed (Conflict)</td><td><strong style="color:#dc3545">{pr.full_conflict_count}</strong></td>
        <td>Partially Shadowed (Conflict)</td><td><strong style="color:#fd7e14">{pr.partial_conflict_count}</strong></td></tr>
    <tr><td>Fully Shadowed (Redundant)</td><td><strong style="color:#198754">{pr.full_redundant_count}</strong></td>
        <td>Partial Overlap (Redundant)</td><td><strong style="color:#0dcaf0">{pr.partial_redundant_count}</strong></td></tr>
    <tr><td>Indeterminate (Unresolved)</td><td><strong style="color:#6c757d">{pr.indeterminate_count}</strong></td>
        <td>Total Findings</td><td><strong>{findings_count}</strong></td></tr>
    <tr><td>Unsupported Objects</td><td>{len(pr.unsupported_objects)}</td>
        <td>Analysis Time</td><td>{pr.elapsed_seconds:.2f}s</td></tr>
  </table>
"""
            if pr.errors:
                html += '  <div class="error-list"><strong>Errors:</strong><ul>'
                for err in pr.errors:
                    html += f"<li>{_esc(err)}</li>"
                html += "</ul></div>"
            html += "</div>\n"
    return html


def _html_findings_detail(package_results: List[PackageResult]) -> str:
    """Render findings grouped by shadowed policy.

    Instead of one collapsible per finding (which explodes when a single
    policy is involved in dozens of relationships), we create one section
    per shadowed policy and display all its shadow relationships as rows
    in a compact table.  The highest-severity finding drives the policy
    section's badge colour.
    """
    from collections import OrderedDict as OD

    total_findings = sum(len(pr.findings) for pr in package_results)

    if not total_findings:
        return "<h2>Findings Detail</h2>\n<p>No findings to display.</p>\n"

    severity_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}

    html = f"<h2>Findings Detail ({total_findings} total)</h2>\n"

    # Group package results by FMG instance
    fmg_groups: OrderedDict[str, List[PackageResult]] = OrderedDict()
    for pr in package_results:
        fmg_groups.setdefault(pr.fmg, []).append(pr)

    package_dom_index = 0
    for fmg_host, prs in fmg_groups.items():
        html += '<div class="fmg-group">\n'
        html += f'<div class="fmg-group-header">FortiManager: {_esc(fmg_host)}</div>\n'

        for pr in prs:
            package_dom_index += 1
            html += '<div class="pkg-group">\n'
            pkg_label = f"{_esc(pr.adom)} / {_esc(pr.package)}"
            finding_count = len(pr.findings)
            html += f'<h4 class="pkg-group-header">{pkg_label} &mdash; {finding_count} finding{"s" if finding_count != 1 else ""}</h4>\n'

            # Show policies with unresolved objects
            unresolved = [p for p in pr.policies if p.has_unresolved]
            if unresolved:
                html += (
                    f'<details class="skipped-list">\n'
                    f'<summary><strong>Policies with Unresolved Objects ({len(unresolved)})</strong>'
                    f' — analyzed with reduced confidence</summary>\n'
                    f'<ul style="margin:4px 0 0 18px">\n'
                )
                for p in unresolved:
                    notes = "; ".join(p.unresolved_notes) if p.unresolved_notes else "unknown"
                    html += (
                        "<li>Policy #{} id={} ({}) &mdash; {}</li>\n".format(
                            _esc(p.seq_num + 1),
                            _esc(p.policyid),
                            _esc(p.name or "unnamed"),
                            _esc(notes),
                        )
                    )
                html += "</ul></details>\n"

            # ── Group findings by shadowed policy ──
            policy_groups = OD()
            for f in pr.findings:
                finding_key = (
                    f.shadowed_section,
                    f.shadowed_seq,
                    f.shadowed_policyid,
                )
                policy_groups.setdefault(finding_key, []).append(f)

            # Sort policy groups by worst severity, then seq
            def _group_sort_key(item):
                pid, findings = item
                best_sev = min(severity_order.get(f.severity_label(), 5) for f in findings)
                seq = findings[0].shadowed_seq
                return (best_sev, seq)

            sorted_groups = sorted(policy_groups.items(), key=_group_sort_key)

            detail_dom_index = 0
            for _shadowed_key, findings in sorted_groups:
                f0 = findings[0]
                worst_sev = min(findings, key=lambda f: severity_order.get(f.severity_label(), 5)).severity_label()
                worst_ftype = min(findings, key=lambda f: severity_order.get(f.severity_label(), 5)).finding_type.value
                max_risk = max(f.risk_score for f in findings)
                n_conflict = sum(1 for f in findings if "conflict" in f.finding_type.value)
                n_redundant = sum(1 for f in findings if "redundant" in f.finding_type.value or "overlap" in f.finding_type.value)
                n_indeterminate = sum(1 for f in findings if "indeterminate" in f.finding_type.value)

                # Summary badges
                badges = _severity_badge(worst_sev) + " " + _finding_type_badge(worst_ftype)

                # Confidence levels present across all findings for this policy
                confidence_levels = sorted(
                    {f.confidence.value for f in findings},
                    key=lambda c: {"low": 0, "medium": 1, "high": 2}.get(c, 3),
                )
                _CONF_STYLES = {
                    "low": "background:#dc354530;color:#dc3545;border:1px solid #dc354560",
                    "medium": "background:#fd7e1430;color:#fd7e14;border:1px solid #fd7e1460",
                    "high": "background:#19875430;color:#198754;border:1px solid #19875460",
                }
                conf_badges = " ".join(
                    f'<span style="display:inline-block;padding:1px 7px;border-radius:4px;'
                    f'font-size:0.75em;font-weight:600;{_CONF_STYLES.get(c, "")}">{c}</span>'
                    for c in confidence_levels
                )

                # Reachability — fully unreachable if ANY finding says so
                any_unreachable = any(f.is_fully_unreachable for f in findings)

                html += f"""<details>
  <summary>
    {badges} {conf_badges}
    &nbsp; Policy #{_esc(f0.shadowed_seq+1)} (id={_esc(f0.shadowed_policyid)}) &mdash; {_esc(f0.shadowed_name or 'unnamed')}{_section_badge(f0.shadowed_section)}
    <span style="float:right;font-size:0.85em;color:var(--muted-text)">
      {len(findings)} relationship{"s" if len(findings) != 1 else ""}
      &bull; risk {max_risk:.1f}
    </span>
  </summary>
  <div class="detail-body">
    <div style="display:flex;gap:16px;flex-wrap:wrap;margin-bottom:10px;">
      <div><strong>Action:</strong> {_esc(f0.shadowed_action)}</div>
      <div><strong>Max Risk:</strong> <span class="risk-score" style="background:{'#dc3545' if max_risk >= 7 else '#fd7e14' if max_risk >= 4 else '#0d6efd'};color:#fff">{max_risk:.1f} / 10.0</span></div>
      <div><strong>Reachability:</strong> {"UNREACHABLE" if any_unreachable else "Partially Reachable"}</div>
    </div>
    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px;font-size:0.85em;">
"""
                if n_conflict:
                    html += f'      <span style="background:#dc354520;color:#dc3545;padding:2px 8px;border-radius:4px;border:1px solid #dc354540">{n_conflict} conflict{"s" if n_conflict != 1 else ""}</span>\n'
                if n_redundant:
                    html += f'      <span style="background:#0dcaf020;color:#0dcaf0;padding:2px 8px;border-radius:4px;border:1px solid #0dcaf040">{n_redundant} redundant</span>\n'
                if n_indeterminate:
                    html += f'      <span style="background:#6c757d20;color:#6c757d;padding:2px 8px;border-radius:4px;border:1px solid #6c757d40">{n_indeterminate} indeterminate</span>\n'

                html += """    </div>
"""

                # ── Relationships table ──
                # Sort: severity desc, risk desc
                sorted_findings = sorted(
                    findings,
                    key=lambda f: (severity_order.get(f.severity_label(), 5), -f.risk_score),
                )

                html += """    <div style="overflow-x:auto;">
    <table class="overlap-table" style="width:100%;">
      <thead><tr>
        <th>Severity</th>
        <th>Type</th>
        <th>Shadowing Rule(s)</th>
        <th>Risk</th>
        <th>Action Conflict</th>
        <th>Confidence</th>
        <th style="width:40px;"></th>
      </tr></thead>
      <tbody>
"""
                for idx, f in enumerate(sorted_findings):
                    sev = f.severity_label()
                    ftype_label = FINDING_TYPE_LABELS.get(f.finding_type.value, f.finding_type.value)
                    ftype_color = FINDING_TYPE_COLORS.get(f.finding_type.value, "#adb5bd")

                    # Pad sections defensively so zip lines up with the ids.
                    sections = f.shadowing_sections or ["local"] * len(f.shadowing_policyids)

                    # Shadowing rules summary
                    if f.is_composite:
                        shad_summary = f"{len(f.shadowing_policyids)} rules (composite)"
                        if any(s in ("global-header", "global-footer") for s in sections):
                            shad_summary += " " + _section_badge("global-header") if "global-header" in sections else ""
                            if "global-footer" in sections:
                                shad_summary += " " + _section_badge("global-footer")
                    else:
                        shad_summary = ", ".join(
                            "#{} id={}{}".format(
                                _esc(s + 1),
                                _esc(pid),
                                _section_badge(sec),
                            )
                            for s, pid, sec in zip(f.shadowing_seqs, f.shadowing_policyids, sections)
                        )
                        # Add name for single-rule findings
                        if len(f.shadowing_names) == 1 and f.shadowing_names[0]:
                            shad_summary += f" ({_esc(f.shadowing_names[0])})"

                    action_html = (
                        f'<span style="color:#dc3545;font-weight:600">DIFFERENT</span> '
                        f'({_esc(f.shadowing_action)} vs {_esc(f.shadowed_action)})'
                        if not f.same_action else
                        f'<span style="color:#198754">Same</span> ({_esc(f.shadowed_action)})'
                    )

                    detail_dom_index += 1
                    detail_id = "detail-{}-{}".format(
                        package_dom_index,
                        detail_dom_index,
                    )

                    html += f"""        <tr style="cursor:pointer" onclick="var el=document.getElementById('{detail_id}');el.style.display=el.style.display==='none'?'table-row':'none'">
          <td>{_severity_badge(sev)}</td>
          <td><span style="color:{ftype_color};font-weight:600;font-size:0.85em">{_esc(ftype_label)}</span></td>
          <td style="font-size:0.9em">{shad_summary}</td>
          <td><span class="risk-score" style="background:{'#dc3545' if f.risk_score >= 7 else '#fd7e14' if f.risk_score >= 4 else '#0d6efd'};color:#fff;font-size:0.8em">{f.risk_score:.1f}</span></td>
          <td style="font-size:0.85em">{action_html}</td>
          <td style="font-size:0.85em">{_esc(f.confidence.value)}</td>
          <td style="font-size:0.8em;color:var(--muted-text)">&#9660;</td>
        </tr>
        <tr id="{detail_id}" style="display:none">
          <td colspan="7" style="padding:8px 12px;border-left:3px solid {ftype_color}">
"""
                    # Inline detail: explanation + overlap table
                    html += f"            {_format_explanation(f.explanation)}\n"

                    if hasattr(f, 'risk_factors') and f.risk_factors:
                        html += f'            <p class="risk-factors" style="font-size:0.85em;margin:6px 0"><strong>Factors:</strong> {_esc(", ".join(f.risk_factors))}</p>\n'

                    # Compact overlap details
                    html += '            <table class="overlap-table" style="font-size:0.85em;margin-top:6px">\n'
                    html += '              <tr><th>Dimension</th><th>Overlap</th></tr>\n'
                    dims = [
                        ("Source Interface", f.srcintf_overlap),
                        ("Dest Interface", f.dstintf_overlap),
                        ("Source Address", f.srcaddr_overlap),
                        ("Dest Address", f.dstaddr_overlap),
                        ("Service", f.service_overlap),
                        ("Schedule", f.schedule_overlap),
                    ]
                    for dim_name, dim_val in dims:
                        html += f"              <tr><td>{dim_name}</td><td>{_esc(dim_val)}</td></tr>\n"
                    html += "            </table>\n"

                    if f.unsupported_notes:
                        html += '            <p style="font-size:0.85em;margin-top:6px"><strong>Unresolved:</strong> ' + _esc("; ".join(f.unsupported_notes)) + '</p>\n'

                    html += """          </td>
        </tr>
"""

                html += """      </tbody>
    </table>
    </div>
  </div>
</details>
"""

            if not sorted_groups:
                html += "<p style='color:#6c757d;font-style:italic'>No findings for this package.</p>\n"

            html += "</div>\n"  # pkg-group

        html += "</div>\n"  # fmg-group

    return html


def _render_cli_candidate_table(
    candidates: List[Dict[str, Any]],
    package_index: int,
) -> str:
    """Render one table of recommended or manually selectable rules."""
    html = """  <div class="remediation-table-wrap">
    <table class="remediation-table">
      <thead>
        <tr>
          <th scope="col" class="checkbox-column">Select</th>
          <th scope="col">Sequence</th>
          <th scope="col">Policy ID</th>
          <th scope="col">Name</th>
          <th scope="col">Selection</th>
          <th scope="col">Analyzer result</th>
          <th scope="col">Action</th>
          <th scope="col">Risk</th>
          <th scope="col">Existing comments</th>
        </tr>
      </thead>
      <tbody>
"""

    for candidate_index, candidate in enumerate(candidates):
        policy = candidate["policy"]
        selection_kind = candidate["selection_kind"]
        checkbox_id = (
            "cli-policy-{}-{}-{}".format(
                package_index,
                selection_kind,
                candidate_index,
            )
        )
        finding_label = "; ".join(candidate["finding_labels"])
        shadowing_ids = ", ".join(candidate["shadowing_policyids"])
        comments = str(policy.comments or "")
        scope = policy.install_scope.describe()
        policy_uuid = str((policy.raw_data or {}).get("uuid", "") or "")
        risk = candidate["max_risk"]
        risk_display = "{:.1f}".format(risk) if risk is not None else "—"
        input_classes = "remediation-policy {}-policy".format(
            selection_kind,
        )
        html += f"""        <tr class="{selection_kind}-candidate">
          <td class="checkbox-column">
            <input type="checkbox"
                   id="{checkbox_id}"
                   class="{input_classes}"
                   autocomplete="off"
                   data-policy-id="{candidate["policy_id"]}"
                   data-sequence="{policy.seq_num + 1}"
                   data-name="{_esc(policy.name or 'unnamed')}"
                   data-comments="{_esc(comments)}"
                   data-selection="{_esc(candidate["script_selection_label"])}"
                   data-selection-kind="{selection_kind}"
                   data-finding="{_esc(finding_label)}"
                   data-shadowing-ids="{_esc(shadowing_ids)}"
                   data-install-scope="{_esc(scope)}"
                   data-uuid="{_esc(policy_uuid)}"
                   aria-label="Select policy ID {candidate["policy_id"]}">
          </td>
          <td>{policy.seq_num + 1}</td>
          <td><label for="{checkbox_id}"><code>{candidate["policy_id"]}</code></label></td>
          <td>{_esc(policy.name or "unnamed")}</td>
          <td><span class="selection-badge {selection_kind}">{_esc(candidate["selection_label"])}</span></td>
          <td>{_esc(finding_label)}</td>
          <td>{_esc(policy.action.value)}</td>
          <td>{risk_display}</td>
          <td class="comments-cell">{_esc(comments or "—")}</td>
        </tr>
"""

    html += """      </tbody>
    </table>
  </div>
"""
    return html


def _html_cli_script_builder(
    package_results: List[PackageResult],
    timestamp: str,
    version: str,
) -> str:
    """Render an offline, package-scoped CLI script builder."""
    prepared = []
    recommended_total = 0
    manual_total = 0
    inherited_total = 0

    for package_index, pr in enumerate(package_results):
        recommended = _collect_disable_candidates(pr)
        manual = _collect_manual_script_candidates(pr, recommended)
        prepared.append((package_index, pr, recommended, manual))
        recommended_total += len(recommended)
        manual_total += len(manual)
        inherited_total += _inherited_script_rule_count(pr)

    available_total = recommended_total + manual_total
    html = f"""<h2>CLI Script Builder</h2>
<div class="card remediation-intro">
  <h3>Build a script to disable policy package rules</h3>
  <p>Choose the rules you want to disable, then preview, copy, or download the CLI script.
     Rules marked <strong>Recommended</strong> were identified by the analyzer as fully
     shadowed. Open <strong>Manual selection</strong> to add any other enabled rule from the
     same policy package.</p>
  <p>The report builds the script in your browser. It does not connect back to FortiManager or
     make any changes by itself. Nothing is selected when the report opens.</p>
  <div class="remediation-warning">
    <strong>Before you run a script:</strong> confirm that you selected the correct policy
    package and review every policy ID and name. A manually selected rule may still process
    traffic, so check the FortiManager package diff and installation preview before installing.
  </div>
  <div class="remediation-stats">
    <span><strong>{recommended_total}</strong> recommended fully shadowed rule{"s" if recommended_total != 1 else ""}</span>
    <span><strong>{manual_total}</strong> other enabled package rule{"s" if manual_total != 1 else ""} available manually</span>
    <span><strong>{inherited_total}</strong> inherited Global Policy Package rule{"s" if inherited_total != 1 else ""} unavailable here</span>
  </div>
"""

    if recommended_total:
        html += """  <div class="remediation-global-controls">
    <label><input type="checkbox" id="remediationSelectAll" autocomplete="off">
      Select all recommended rules in this report
    </label>
    <button type="button" id="remediationClearAll" class="secondary-button">Clear all</button>
    <span id="remediationGlobalStatus" class="selection-status">0 selected</span>
  </div>
"""
    elif available_total:
        html += """  <p class="remediation-empty"><strong>No rules were recommended
     automatically.</strong> You can still open Manual selection under a policy package and
     choose individual rules.</p>
  <div class="remediation-global-controls">
    <button type="button" id="remediationClearAll" class="secondary-button">Clear all</button>
    <span id="remediationGlobalStatus" class="selection-status">0 selected</span>
  </div>
"""
    else:
        html += """  <p class="remediation-empty"><strong>No rules are available for this
     builder.</strong> It only lists enabled rules defined in the policy package with a valid,
     unique policy ID.</p>
"""
    if available_total:
        html += """  <p class="remediation-note">The report creates a separate script for each
     policy package. Download each script from its package card.</p>
"""
    html += "</div>\n"

    for package_index, pr, recommended, manual in prepared:
        if not recommended and not manual:
            continue

        builder_id = f"cli-builder-{package_index}"
        html += f"""<div class="card cli-builder"
     id="{builder_id}"
     data-fmg="{_esc(pr.fmg)}"
     data-adom="{_esc(pr.adom)}"
     data-package="{_esc(pr.package)}"
     data-timestamp="{_esc(timestamp)}"
     data-version="{_esc(version)}">
  <h3>{_esc(pr.fmg)} / {_esc(pr.adom)} / {_esc(pr.package)}</h3>
  <p class="remediation-target"><strong>Where to run this script:</strong> create or import a
     <strong>CLI Script</strong>, select <strong>Policy Package or ADOM Database</strong>, and
     choose <strong>{_esc(pr.package)}</strong>. Do not select
     <em>Remote FortiGate Directly</em>.</p>
  <div class="remediation-package-controls">
"""
        if recommended:
            html += f"""    <label><input type="checkbox" class="package-select-all" autocomplete="off">
      Select all {len(recommended)} recommended rule{"s" if len(recommended) != 1 else ""} in this package
    </label>
"""
        html += """    <span class="selection-status">0 selected</span>
  </div>
"""

        if recommended:
            html += """  <h4>Recommended fully shadowed rules</h4>
  <p class="remediation-note">The analyzer recommends these rules because it found them fully
     shadowed. Select them individually or use the checkbox above.</p>
"""
            html += _render_cli_candidate_table(
                recommended,
                package_index,
            )

        if manual:
            manual_summary = (
                "Manual selection: choose from {} other enabled rule{}".format(
                    len(manual),
                    "s" if len(manual) != 1 else "",
                )
            )
            html += f"""  <details class="manual-policy-picker">
    <summary data-base-label="{_esc(manual_summary)}">{_esc(manual_summary)}</summary>
    <p>These rules were not recommended automatically. Select a rule here only when you
       intend to disable it and have reviewed the effect.</p>
"""
            html += _render_cli_candidate_table(
                manual,
                package_index,
            )
            html += "  </details>\n"

        html += """  <div class="remediation-actions">
    <button type="button" class="preview-cli" disabled>Preview script</button>
    <button type="button" class="copy-cli secondary-button" disabled>Copy script</button>
    <button type="button" class="download-cli" disabled>Download .txt</button>
    <span class="copy-status" aria-live="polite"></span>
  </div>
  <textarea class="cli-preview" readonly hidden
            aria-label="Generated FortiManager CLI script preview"></textarea>
  <p class="remediation-note">Generate a fresh report before running the script. Check each
     policy ID, name, and UUID in the package diff. If your FortiManager uses workspace or
     workflow mode, follow its normal lock, approval, and install process.</p>
</div>
"""

    return html


def _html_methodology() -> str:
    return '''<h2>Methodology</h2>
<div class="methodology">
  <p><strong>Policy Shadow &amp; Overlap Analysis</strong> evaluates each firewall rule against all
  higher-priority rules (rules appearing earlier in the policy order). Under first-match
  evaluation semantics, the firewall processes rules sequentially from top to bottom &mdash;
  when a rule matches, its action is applied and no further rules are checked. A lower-priority
  rule is &quot;shadowed&quot; when a higher-priority rule matches the same or broader traffic.</p>

  <h3 style="margin-top:14px;font-size:1em">Finding Classifications</h3>
  <ul style="margin:8px 0 0 20px">
    <li><strong>Fully Shadowed (Conflict)</strong> &mdash; A higher-priority rule with a <em>different</em>
        action completely covers the shadowed rule&rsquo;s match space. The shadowed rule is
        <strong>unreachable</strong> and will never match any traffic. This is the highest-risk
        finding &mdash; it may indicate a security gap where intended deny rules are being
        overridden by allow rules, or vice versa.
        <em>(Industry terms: contradictory shadow, conflicting rule)</em></li>
    <li><strong>Partially Shadowed (Conflict)</strong> &mdash; A higher-priority rule with a
        <em>different</em> action covers part of the shadowed rule&rsquo;s traffic. Some traffic
        reaches a different verdict than intended.
        <em>(Industry terms: partial shadow, partially covered rule)</em></li>
    <li><strong>Fully Shadowed (Redundant)</strong> &mdash; A higher-priority rule with the
        <em>same</em> action completely covers the shadowed rule. The rule is redundant &mdash;
        removing it would not change the security posture, but would simplify the rule base.
        <em>(Industry terms: redundant rule, covered rule)</em></li>
    <li><strong>Partially Overlapping (Redundant)</strong> &mdash; A higher-priority rule with
        the <em>same</em> action covers part of the shadowed rule&rsquo;s traffic. The overlap
        adds complexity but is not a security issue.
        <em>(Industry terms: partial redundancy, overlapping rule)</em></li>
    <li><strong>Indeterminate (Unresolved Objects)</strong> &mdash; Overlap was detected but
        unresolved objects (FQDNs, dynamic groups, geographic addresses, etc.) prevent
        definitive classification. Manual review recommended.
        <em>(Industry terms: inconclusive analysis)</em></li>
  </ul>

  <h3 style="margin-top:14px;font-size:1em">Risk Scoring</h3>
  <p>Each finding receives a risk score (0&ndash;10) based on weighted factors:</p>
  <ul style="margin:6px 0 0 20px">
    <li><strong>Interface overlaps</strong> are common and expected between policies &mdash;
        they carry minimal weight in the risk score.</li>
    <li><strong>Address object overlaps</strong> (source/destination) are more significant
        and carry higher weight.</li>
    <li><strong>Combined address + service + schedule overlaps</strong> represent the most
        complete shadowing and carry the highest weight.</li>
    <li><strong>Address breadth</strong> amplifies risk: a rule affecting all hosts (0.0.0.0/0)
        scores higher than one affecting a single host.</li>
    <li><strong>Action conflicts</strong> (deny shadowed by allow, or vice versa) receive
        a multiplier due to their security implications.</li>
  </ul>

  <h3 style="margin-top:14px;font-size:1em">Dimensions Compared</h3>
  <p style="margin-top:6px">Six match dimensions are evaluated: Source Interface, Destination Interface,
  Source Address, Destination Address, Service (protocol/ports), and Schedule. For full shadow
  detection, all six dimensions must be covered by the higher-priority rule(s). A pairwise full
  finding also requires the higher-priority rule&rsquo;s install scope to contain every target
  in the shadowed rule&rsquo;s install scope.</p>

  <h3 style="margin-top:14px;font-size:1em">Compliance Context</h3>
  <p style="margin-top:6px">Regular rule base review aligns with PCI DSS Requirement 1.1.7
  (semi-annual firewall rule review), CIS Controls (configuration review), NIST SP 800-41
  (firewall policy guidelines), and ISO 27001 (access control management).</p>
</div>
'''


def _html_limitations() -> str:
    return """<h2>Limitations</h2>
<div class="limitations">
  <ul style="margin-left:20px">
    <li>FQDN-based address objects cannot be fully resolved to IP ranges and are marked indeterminate.</li>
    <li>Dynamic address groups and SDN connectors are not expanded.</li>
    <li>Geographic/country-based address objects are approximated or marked indeterminate.</li>
    <li>Schedule overlap analysis uses simplified time comparisons.</li>
    <li>Internet-service (ISDB) references are labeled from the FortiGuard catalog, but their
        full address/port datasets are not expanded; affected policies remain indeterminate.</li>
    <li>Recognized direct identity, SGT, ZTNA, IPv6, NGFW application/URL, dynamic-service,
        ToS, expiry, VIP-match, and Internet-service selectors are marked unresolved. Rules
        using them are not recommended automatically, but may be added through Manual
        selection. Version-specific response fields may still vary.</li>
    <li>Ordinary post-selection security profile assignments (AV, Web Filter, App Control
        profiles, IPS, and similar) are captured and displayed for context; they are not modeled
        as policy-selection dimensions.</li>
    <li>VIP/DNAT destination translations are resolved where possible but complex scenarios may be missed.</li>
    <li>Policy sets with install-scope restrictions are compared only when scopes overlap.</li>
    <li>Composite-union shadow detection is heuristic. Composite findings are not recommended
        automatically, but rules defined in the selected package remain available through
        Manual selection.</li>
    <li>Generated scripts can edit only rules defined directly in the selected policy package.
        Inherited Global Header/Footer rules require a separate Global Database workflow.</li>
    <li>Central SNAT/DNAT policies are not included in the analysis.</li>
  </ul>
</div>
"""


def _html_footer() -> str:
    return """</div>
<script>
(function() {
  var toggle = document.getElementById('darkToggle');
  var body = document.body;
  // Restore saved preference
  try {
    if (localStorage.getItem('fmg-dark-mode') === 'true') {
      body.classList.add('dark');
      toggle.innerHTML = '&#9788;';  // sun
    }
  } catch (error) {
    // Keep the default theme when file:// storage is unavailable.
  }
  toggle.addEventListener('click', function() {
    body.classList.toggle('dark');
    var isDark = body.classList.contains('dark');
    try {
      localStorage.setItem('fmg-dark-mode', isDark);
    } catch (error) {
      // Theme still toggles for this page even if it cannot be persisted.
    }
    toggle.innerHTML = isDark ? '&#9788;' : '&#9790;';  // sun : moon
    document.documentElement.style.background = isDark ? '#0d1117' : '';
  });
})();

(function() {
  var builders = Array.prototype.slice.call(document.querySelectorAll('.cli-builder'));
  if (!builders.length) {
    return;
  }

  var globalSelect = document.getElementById('remediationSelectAll');
  var clearAll = document.getElementById('remediationClearAll');
  var globalStatus = document.getElementById('remediationGlobalStatus');

  function safeMetadata(value) {
    return String(value || '')
      .replace(/[\\u0000-\\u001f\\u007f-\\u009f]+/g, ' ')
      .replace(/[\\u00ad\\u061c\\u200b-\\u200f\\u202a-\\u202e\\u2060-\\u206f\\ufeff]+/g, '')
      .replace(/\\s+/g, ' ')
      .trim();
  }

  function selectedPolicies(builder) {
    return Array.prototype.slice.call(
      builder.querySelectorAll('.remediation-policy:checked')
    ).sort(function(a, b) {
      var seqDiff = Number(a.dataset.sequence) - Number(b.dataset.sequence);
      if (seqDiff) {
        return seqDiff;
      }
      return Number(a.dataset.policyId) - Number(b.dataset.policyId);
    });
  }

  function buildCliScript(builder) {
    var selected = selectedPolicies(builder);
    var manualCount = selected.filter(function(input) {
      return input.dataset.selectionKind === 'manual';
    }).length;
    var lines = [
      '# Generated by FMG Policy Shadow Analyzer ' + safeMetadata(builder.dataset.version),
      '# Report timestamp: ' + safeMetadata(builder.dataset.timestamp),
      '# FortiManager: ' + safeMetadata(builder.dataset.fmg),
      '# ADOM: ' + safeMetadata(builder.dataset.adom),
      '# Policy package: ' + safeMetadata(builder.dataset.package),
      '# Script type: CLI',
      '# Run script on: Policy Package or ADOM Database',
      '# WARNING: Run only against the exact package above; do not run directly on a FortiGate.',
      '# WARNING: Point-in-time output; regenerate immediately before execution.',
      '# WARNING: Verify every policy ID, name, and UUID in the package diff before installation.'
    ];
    if (manualCount) {
      lines.push(
        '# WARNING: ' + manualCount
        + (manualCount === 1 ? ' rule was' : ' rules were')
        + ' added manually and may still process traffic.'
      );
    }
    lines.push('');
    lines.push('config firewall policy');

    selected.forEach(function(input) {
      var policyId = safeMetadata(input.dataset.policyId);
      if (
        !/^[1-9][0-9]*$/.test(policyId)
        || Number(policyId) > 1071741824
      ) {
        return;
      }
      lines.push('# ------------------------------------------------------------------');
      lines.push('# Policy ID: ' + policyId);
      lines.push('# Sequence at analysis time: ' + safeMetadata(input.dataset.sequence));
      lines.push('# Name: ' + (safeMetadata(input.dataset.name) || 'unnamed'));
      lines.push('# Selection: ' + safeMetadata(input.dataset.selection));
      if (safeMetadata(input.dataset.uuid)) {
        lines.push('# UUID at analysis time: ' + safeMetadata(input.dataset.uuid));
      }
      lines.push('# Existing comments: ' + (safeMetadata(input.dataset.comments) || '(none)'));
      lines.push('# Analyzer result: ' + safeMetadata(input.dataset.finding));
      if (safeMetadata(input.dataset.shadowingIds)) {
        lines.push('# Shadowing policy ID(s): ' + safeMetadata(input.dataset.shadowingIds));
      }
      lines.push('# Installation scope: ' + safeMetadata(input.dataset.installScope));
      lines.push('edit ' + policyId);
      lines.push('    set status disable');
      lines.push('next');
    });

    lines.push('end');
    lines.push('');
    return lines.join('\\n');
  }

  function safeFilenamePart(value) {
    var part = safeMetadata(value)
      .replace(/[^A-Za-z0-9._-]+/g, '_')
      .replace(/^_+|_+$/g, '');
    return part || 'unknown';
  }

  function clearRestoredSelections() {
    document.querySelectorAll(
      '.remediation-policy, .package-select-all, #remediationSelectAll'
    ).forEach(function(input) {
      input.checked = false;
      input.indeterminate = false;
    });
  }

  function scriptFilename(builder) {
    return [
      'disable-selected-policies',
      safeFilenamePart(builder.dataset.fmg),
      safeFilenamePart(builder.dataset.adom),
      safeFilenamePart(builder.dataset.package)
    ].join('_') + '.txt';
  }

  function syncGlobalStatus() {
    var all = Array.prototype.slice.call(
      document.querySelectorAll('.remediation-policy')
    );
    var recommended = Array.prototype.slice.call(
      document.querySelectorAll('.recommended-policy')
    );
    var selected = all.filter(function(input) { return input.checked; });
    var selectedRecommended = selected.filter(function(input) {
      return input.dataset.selectionKind === 'recommended';
    }).length;
    var selectedManual = selected.length - selectedRecommended;
    if (globalStatus) {
      globalStatus.textContent = selected.length + ' selected';
      if (selected.length) {
        globalStatus.textContent += (
          ' (' + selectedRecommended + ' recommended, '
          + selectedManual + ' manual)'
        );
      }
    }
    if (globalSelect) {
      globalSelect.checked = (
        recommended.length > 0
        && selectedRecommended === recommended.length
      );
      globalSelect.indeterminate = (
        selectedRecommended > 0
        && selectedRecommended < recommended.length
      );
    }
  }

  function syncBuilder(builder) {
    var recommended = Array.prototype.slice.call(
      builder.querySelectorAll('.recommended-policy')
    );
    var selected = selectedPolicies(builder);
    var count = selected.length;
    var selectedRecommended = selected.filter(function(input) {
      return input.dataset.selectionKind === 'recommended';
    }).length;
    var selectedManual = count - selectedRecommended;
    var packageToggle = builder.querySelector('.package-select-all');
    var status = builder.querySelector('.selection-status');
    var preview = builder.querySelector('.cli-preview');
    var manualSummary = builder.querySelector(
      '.manual-policy-picker summary'
    );

    status.textContent = count + ' selected';
    if (count) {
      status.textContent += (
        ' (' + selectedRecommended + ' recommended, '
        + selectedManual + ' manual)'
      );
    }
    if (packageToggle) {
      packageToggle.checked = (
        recommended.length > 0
        && selectedRecommended === recommended.length
      );
      packageToggle.indeterminate = (
        selectedRecommended > 0
        && selectedRecommended < recommended.length
      );
    }
    if (manualSummary) {
      manualSummary.textContent = manualSummary.dataset.baseLabel;
      if (selectedManual) {
        manualSummary.textContent += ' — ' + selectedManual + ' selected';
      }
    }

    Array.prototype.forEach.call(
      builder.querySelectorAll('.preview-cli, .copy-cli, .download-cli'),
      function(button) { button.disabled = count === 0; }
    );

    if (!preview.hidden && count > 0) {
      preview.value = buildCliScript(builder);
    } else if (count === 0) {
      preview.value = '';
      preview.hidden = true;
      builder.querySelector('.preview-cli').textContent = 'Preview script';
    }
    syncGlobalStatus();
  }

  function copyText(text) {
    if (navigator.clipboard && window.isSecureContext) {
      return navigator.clipboard.writeText(text);
    }
    return new Promise(function(resolve, reject) {
      var helper = document.createElement('textarea');
      helper.value = text;
      helper.setAttribute('readonly', '');
      helper.style.position = 'fixed';
      helper.style.opacity = '0';
      document.body.appendChild(helper);
      helper.select();
      try {
        if (!document.execCommand('copy')) {
          throw new Error('Copy command was rejected');
        }
        resolve();
      } catch (error) {
        reject(error);
      } finally {
        document.body.removeChild(helper);
      }
    });
  }

  clearRestoredSelections();

  builders.forEach(function(builder) {
    var packageToggle = builder.querySelector('.package-select-all');
    var previewButton = builder.querySelector('.preview-cli');
    var copyButton = builder.querySelector('.copy-cli');
    var downloadButton = builder.querySelector('.download-cli');
    var preview = builder.querySelector('.cli-preview');
    var copyStatus = builder.querySelector('.copy-status');

    builder.querySelectorAll('.remediation-policy').forEach(function(input) {
      input.addEventListener('change', function() {
        copyStatus.textContent = '';
        syncBuilder(builder);
      });
    });

    if (packageToggle) {
      packageToggle.addEventListener('change', function() {
        builder.querySelectorAll('.recommended-policy').forEach(function(input) {
          input.checked = packageToggle.checked;
        });
        copyStatus.textContent = '';
        syncBuilder(builder);
      });
    }

    previewButton.addEventListener('click', function() {
      preview.hidden = !preview.hidden;
      if (!preview.hidden) {
        preview.value = buildCliScript(builder);
        previewButton.textContent = 'Hide preview';
      } else {
        previewButton.textContent = 'Preview script';
      }
    });

    copyButton.addEventListener('click', function() {
      var script = buildCliScript(builder);
      preview.value = script;
      copyText(script).then(function() {
        copyStatus.textContent = 'Copied.';
      }).catch(function() {
        copyStatus.textContent = 'Copy failed; use the preview and copy manually.';
        preview.hidden = false;
        previewButton.textContent = 'Hide preview';
      });
    });

    downloadButton.addEventListener('click', function() {
      var blob = new Blob([buildCliScript(builder)], {
        type: 'text/plain;charset=utf-8'
      });
      var url = URL.createObjectURL(blob);
      var link = document.createElement('a');
      link.href = url;
      link.download = scriptFilename(builder);
      document.body.appendChild(link);
      link.click();
      document.body.removeChild(link);
      window.setTimeout(function() { URL.revokeObjectURL(url); }, 0);
      copyStatus.textContent = 'Downloaded ' + link.download + '.';
    });

    syncBuilder(builder);
  });

  if (globalSelect) {
    globalSelect.addEventListener('change', function() {
      document.querySelectorAll('.recommended-policy').forEach(function(input) {
        input.checked = globalSelect.checked;
      });
      builders.forEach(syncBuilder);
    });
  }

  if (clearAll) {
    clearAll.addEventListener('click', function() {
      document.querySelectorAll('.remediation-policy').forEach(function(input) {
        input.checked = false;
      });
      builders.forEach(syncBuilder);
    });
  }

  window.addEventListener('pageshow', function() {
    clearRestoredSelections();
    builders.forEach(syncBuilder);
  });
})();
</script>
</body>
</html>
"""


# ===================================================================
# Excel Report
# ===================================================================

def generate_excel_report(run_result: RunResult, output_path: str) -> Optional[str]:
    """Generate an Excel (.xlsx) report with multiple sheets.

    Returns the output file path, or None if openpyxl is unavailable.
    """
    if not HAS_OPENPYXL:
        print("[WARNING] openpyxl is not installed. Skipping Excel report generation.")
        print("  Install with: pip install openpyxl")
        return None

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _excel_summary_sheet(wb, run_result)
    _excel_findings_sheet(wb, run_result)
    _excel_packages_sheet(wb, run_result)
    _excel_unsupported_sheet(wb, run_result)
    _excel_policy_inventory_sheet(wb, run_result)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


# -- Style helpers --

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11) if HAS_OPENPYXL else None
_HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid") if HAS_OPENPYXL else None
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True) if HAS_OPENPYXL else None
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top") if HAS_OPENPYXL else None
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
) if HAS_OPENPYXL else None

_SEVERITY_FILLS: Dict[str, Any] = {}
if HAS_OPENPYXL:
    _SEVERITY_FILLS = {
        "CRITICAL": PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid"),
        "HIGH": PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid"),
        "MEDIUM": PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
        "LOW": PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid"),
        "INFO": PatternFill(start_color="ADB5BD", end_color="ADB5BD", fill_type="solid"),
    }


def _apply_header_row(ws, headers: List[str], widths: Optional[List[int]]= None):
    """Write header row with formatting."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    # Column widths
    if widths:
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    else:
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    if headers:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}1"


def _write_row(ws, row_num: int, values: list, wrap_cols: Optional[Set[int]]= None):
    """Write a data row."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.border = _THIN_BORDER
        if wrap_cols and col_idx in wrap_cols:
            cell.alignment = _WRAP_ALIGNMENT


# -- Sheet builders --

def _excel_summary_sheet(wb, run_result: RunResult):
    ws = wb.create_sheet("Summary")
    counts = run_result.summary_counts()

    # Metadata section
    meta = [
        ("Tool Version", run_result.tool_version),
        ("Run Timestamp", run_result.run_timestamp),
        ("Elapsed Seconds", run_result.elapsed_seconds),
        ("", ""),
        ("SUMMARY", ""),
        ("Packages Analyzed", counts.get("packages_analyzed", 0)),
        ("Total Policies", counts.get("total_policies", 0)),
        ("Total Findings", counts.get("total_findings", 0)),
        ("", ""),
        ("FINDINGS BY TYPE", ""),
        ("Full Conflict Shadow", counts.get("full_conflict_shadow", 0)),
        ("Partial Conflict Shadow", counts.get("partial_conflict_shadow", 0)),
        ("Full Redundant Coverage", counts.get("full_redundant_coverage", 0)),
        ("Partial Redundant Overlap", counts.get("partial_redundant_overlap", 0)),
        ("Indeterminate", counts.get("indeterminate", 0)),
        ("Errors", counts.get("errors", 0)),
    ]

    bold_font = Font(bold=True, size=11)
    section_font = Font(bold=True, size=12, color="1A1A2E")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30

    for row_idx, (label, value) in enumerate(meta, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if label in ("SUMMARY", "FINDINGS BY TYPE"):
            cell_a.font = section_font
        else:
            cell_a.font = bold_font

    # Per-package stats table below
    start_row = len(meta) + 2
    ws.cell(row=start_row, column=1, value="PER-PACKAGE STATS").font = section_font
    start_row += 1

    pkg_headers = ["FMG", "ADOM", "Package", "Total Policies", "Effective",
                   "Full Conflict", "Partial Conflict", "Full Redundant",
                   "Partial Redundant", "Indeterminate", "Total Findings", "Time (s)"]
    for col_idx, h in enumerate(pkg_headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    for i, pr in enumerate(run_result.package_results):
        row = start_row + 1 + i
        vals = [pr.fmg, pr.adom, pr.package, pr.total_policies, pr.effective_policies,
                pr.full_conflict_count, pr.partial_conflict_count, pr.full_redundant_count,
                pr.partial_redundant_count, pr.indeterminate_count, len(pr.findings),
                round(pr.elapsed_seconds, 2)]
        for col_idx, v in enumerate(vals, 1):
            ws.cell(row=row, column=col_idx, value=v)


def _excel_findings_sheet(wb, run_result: RunResult):
    ws = wb.create_sheet("Findings")

    headers = [
        "FMG", "ADOM", "Package",
        "Shadowed Policy ID", "Shadowed Name", "Shadowed Seq", "Shadowed Section",
        "Finding Type", "Severity", "Is Composite",
        "Shadowing Policy IDs", "Shadowing Names", "Shadowing Seqs", "Shadowing Sections",
        "SrcIntf Overlap", "DstIntf Overlap",
        "SrcAddr Overlap", "DstAddr Overlap",
        "Service Overlap", "Schedule Overlap",
        "Fully Unreachable", "Residual Description",
        "Shadowed Action", "Shadowing Action", "Same Action",
        "Confidence", "Unsupported Notes", "Explanation",
    ]
    widths = [
        14, 14, 18,
        16, 20, 12, 14,
        28, 12, 12,
        22, 22, 18, 18,
        20, 20,
        25, 25,
        20, 18,
        14, 30,
        14, 14, 12,
        12, 30, 50,
    ]
    wrap_cols = {17, 18, 22, 27, 28}  # long text columns

    _apply_header_row(ws, headers, widths)

    row_num = 2
    for pr in run_result.package_results:
        for f in pr.findings:
            d = f.to_dict()
            values = [
                d["fmg"], d["adom"], d["package"],
                d["shadowed_policyid"], d["shadowed_name"], d["shadowed_seq"],
                _SECTION_LABELS.get(d["shadowed_section"], d["shadowed_section"]),
                FINDING_TYPE_LABELS.get(d["finding_type"], d["finding_type"]),
                d["severity"], d["is_composite"],
                ", ".join(str(x) for x in d["shadowing_policyids"]),
                ", ".join(str(x) for x in d["shadowing_names"]),
                ", ".join(str(x) for x in d["shadowing_seqs"]),
                ", ".join(
                    _SECTION_LABELS.get(s, s) for s in d.get("shadowing_sections", [])
                ),
                d["srcintf_overlap"], d["dstintf_overlap"],
                d["srcaddr_overlap"], d["dstaddr_overlap"],
                d["service_overlap"], d["schedule_overlap"],
                d["is_fully_unreachable"], d["residual_description"],
                d["shadowed_action"], d["shadowing_action"], d["same_action"],
                d["confidence"],
                "; ".join(d["unsupported_notes"]),
                d["explanation"],
            ]
            _write_row(ws, row_num, values, wrap_cols)

            # Color the severity cell
            sev = d["severity"]
            sev_cell = ws.cell(row=row_num, column=9)
            if sev in _SEVERITY_FILLS:
                sev_cell.fill = _SEVERITY_FILLS[sev]
                if sev in ("CRITICAL", "HIGH", "LOW"):
                    sev_cell.font = Font(bold=True, color="FFFFFF")
                else:
                    sev_cell.font = Font(bold=True)

            row_num += 1


def _excel_packages_sheet(wb, run_result: RunResult):
    ws = wb.create_sheet("Packages")

    headers = [
        "FMG", "ADOM", "Package",
        "Total Policies", "Global Header", "Local", "Global Footer",
        "Effective Policies",
        "Full Conflict", "Partial Conflict",
        "Full Redundant", "Partial Redundant",
        "Indeterminate", "Total Findings",
        "Unsupported Objects", "Errors", "Time (s)",
    ]
    widths = [14, 14, 20, 14, 14, 10, 14, 16, 14, 16, 14, 16, 14, 14, 18, 10, 10]

    _apply_header_row(ws, headers, widths)

    for i, pr in enumerate(run_result.package_results):
        row = i + 2
        local_count = (
            pr.total_policies - pr.global_header_policies - pr.global_footer_policies
        )
        values = [
            pr.fmg, pr.adom, pr.package,
            pr.total_policies, pr.global_header_policies, local_count,
            pr.global_footer_policies,
            pr.effective_policies,
            pr.full_conflict_count, pr.partial_conflict_count,
            pr.full_redundant_count, pr.partial_redundant_count,
            pr.indeterminate_count, len(pr.findings),
            len(pr.unsupported_objects), len(pr.errors),
            round(pr.elapsed_seconds, 2),
        ]
        _write_row(ws, row, values)


def _excel_unsupported_sheet(wb, run_result: RunResult):
    ws = wb.create_sheet("Unsupported Objects")

    headers = ["FMG", "ADOM", "Package", "Object"]
    widths = [14, 14, 20, 50]

    _apply_header_row(ws, headers, widths)

    row_num = 2
    for pr in run_result.package_results:
        for obj in pr.unsupported_objects:
            _write_row(ws, row_num, [pr.fmg, pr.adom, pr.package, obj], {4})
            row_num += 1

    if row_num == 2:
        ws.cell(row=2, column=1, value="No unsupported objects found.")


def _excel_policy_inventory_sheet(wb, run_result: RunResult):
    ws = wb.create_sheet("Policy Inventory")

    headers = [
        "FMG", "ADOM", "Package",
        "Policy ID", "Name", "Seq #", "Section",
        "Action", "Status",
        "Source Interface", "Dest Interface",
        "SrcAddr Objects", "SrcAddr Resolved",
        "DstAddr Objects", "DstAddr Resolved",
        "Service Objects", "Service Resolved",
        "Schedule",
        "Comments", "Has Unresolved", "Security Profiles",
    ]
    widths = [14, 14, 18, 10, 20, 8, 14, 10, 10, 16, 16, 25, 25, 25, 25, 20, 20, 14, 30, 14, 35]
    wrap_cols = {12, 13, 14, 15, 19, 21}

    _apply_header_row(ws, headers, widths)

    row_num = 2
    for pr in run_result.package_results:
        for p in pr.policies:
            raw = p.raw_data or {}
            values = [
                p.fmg, p.adom, p.package,
                p.policyid, p.name, p.seq_num,
                _SECTION_LABELS.get(p.policy_section, p.policy_section),
                p.action.value if p.action else "",
                p.status,
                p.srcintf.describe() if p.srcintf else "",
                p.dstintf.describe() if p.dstintf else "",
                ", ".join(raw.get("_raw_srcaddr", [])),
                p.srcaddr.describe() if p.srcaddr else "",
                ", ".join(raw.get("_raw_dstaddr", [])),
                p.dstaddr.describe() if p.dstaddr else "",
                ", ".join(raw.get("_raw_service", [])),
                p.service.describe() if p.service else "",
                p.schedule.describe() if p.schedule else "",
                p.comments, p.has_unresolved,
                ', '.join(f'{k}={v}' for k, v in sorted(p.security_profiles.items())) if p.security_profiles else '',
            ]
            _write_row(ws, row_num, values, wrap_cols)
            row_num += 1


# ===================================================================
# Convenience: generate all reports
# ===================================================================

def generate_all_reports(
    run_result: RunResult,
    output_dir: str,
    formats: Optional[List[str]]= None,
) -> Dict[str, str]:
    """Generate reports in all requested formats.

    Args:
        run_result: The analysis results.
        output_dir: Directory for output files.
        formats: List of format strings: "json", "html", "excel"/"xlsx".
                 Defaults to all formats.

    Returns:
        Dict mapping format name to output file path.
    """
    if formats is None:
        formats = ["json", "html", "excel"]

    os.makedirs(output_dir, exist_ok=True)

    timestamp_slug = (run_result.run_timestamp or datetime.now().isoformat()).replace(":", "-").replace(" ", "_")[:19]
    base = f"shadow_report_{timestamp_slug}"

    results: Dict[str, str] = {}

    for fmt in formats:
        fmt_lower = fmt.lower()
        if fmt_lower == "json":
            path = os.path.join(output_dir, f"{base}.json")
            generate_json_report(run_result, path)
            results["json"] = path

        elif fmt_lower == "html":
            path = os.path.join(output_dir, f"{base}.html")
            generate_html_report(run_result, path)
            results["html"] = path

        elif fmt_lower in ("excel", "xlsx"):
            path = os.path.join(output_dir, f"{base}.xlsx")
            result = generate_excel_report(run_result, path)
            if result:
                results["excel"] = path

        else:
            print(f"[WARNING] Unknown report format: {fmt}")

    return results
